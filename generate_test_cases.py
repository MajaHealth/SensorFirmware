#!/usr/bin/env python3
"""
Generate test cases for SPI_DEV_servise and power_control_service in Excel format
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def create_test_cases_excel():
    # Create workbook and sheets
    wb = openpyxl.Workbook()

    # Remove default sheet and create two sheets
    wb.remove(wb.active)
    ws_spi = wb.create_sheet("SPI_DEV_servise Tests")
    ws_pwr = wb.create_sheet("power_control_service Tests")

    # Define column headers
    headers = ["Test ID", "Area", "Objective", "Linked IDs (Hazard/SRS/Task)",
               "Preconditions", "Steps", "Expected", "Dataset", "Priority",
               "Result", "Defect ID"]

    # Style definitions
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # SPI_DEV_servise test cases
    spi_tests = [
        {
            "Test ID": "SPI-001",
            "Area": "MAX30009 Communication",
            "Objective": "Verify MAX30009 TCP server accepts connections on port 30009",
            "Linked IDs (Hazard/SRS/Task)": "SRS-SPI-001",
            "Preconditions": "SPI_DEV_servise running, network available",
            "Steps": "1. Start SPI_DEV_servise\n2. Connect to localhost:30009 via telnet\n3. Verify connection established",
            "Expected": "TCP connection established successfully, server ready to receive JSON commands",
            "Dataset": "N/A",
            "Priority": "High",
            "Result": "",
            "Defect ID": ""
        },
        {
            "Test ID": "SPI-002",
            "Area": "MAX30009 Data Acquisition",
            "Objective": "Verify MAX30009 bioimpedance data streaming with I/Q samples",
            "Linked IDs (Hazard/SRS/Task)": "SRS-SPI-002, HAZ-001",
            "Preconditions": "MAX30009 initialized, sensor configured, hardware connected",
            "Steps": "1. Send JSON start command to port 30009\n2. Configure frequency (25Hz-450kHz) and current (64μA-1.28mA)\n3. Start data acquisition\n4. Verify I/Q data FIFO buffering (max 30,000 samples)",
            "Expected": "I/Q data stream received continuously, FIFO does not overflow, data format correct",
            "Dataset": "freq=100kHz, current=500μA",
            "Priority": "Critical",
            "Result": "",
            "Defect ID": ""
        },
        {
            "Test ID": "SPI-003",
            "Area": "MAX30009 Sync Marks",
            "Objective": "Verify sync marks injected every 1 second with magic number -99999",
            "Linked IDs (Hazard/SRS/Task)": "SRS-SPI-003",
            "Preconditions": "MAX30009 streaming data",
            "Steps": "1. Start MAX30009 data acquisition\n2. Monitor data stream for 60 seconds\n3. Count sync marks\n4. Verify magic number -99999 at sync positions",
            "Expected": "Exactly 60 sync marks with value -99999, evenly spaced at 1-second intervals",
            "Dataset": "60 second capture",
            "Priority": "High",
            "Result": "",
            "Defect ID": ""
        },
        {
            "Test ID": "SPI-004",
            "Area": "MAX30009 Calibration",
            "Objective": "Verify automated calibration system across 17 frequency points × 5 current levels",
            "Linked IDs (Hazard/SRS/Task)": "SRS-SPI-004",
            "Preconditions": "Calibration reference load connected, calib/ directory exists",
            "Steps": "1. Send calibration start command\n2. Monitor calibration progress (85 total points)\n3. Verify JSON files created in calib/ with format {current_index}_{freq_index}.json\n4. Validate calibration data structure",
            "Expected": "85 calibration files created successfully, data within expected ranges, no missing points",
            "Dataset": "Full calibration sweep",
            "Priority": "High",
            "Result": "",
            "Defect ID": ""
        },
        {
            "Test ID": "SPI-005",
            "Area": "ADS1293 Communication",
            "Objective": "Verify ADS1293 TCP server accepts connections on port 1293",
            "Linked IDs (Hazard/SRS/Task)": "SRS-SPI-005",
            "Preconditions": "SPI_DEV_servise running, network available",
            "Steps": "1. Start SPI_DEV_servise\n2. Connect to localhost:1293 via telnet\n3. Verify connection established",
            "Expected": "TCP connection established successfully, server ready to receive JSON commands",
            "Dataset": "N/A",
            "Priority": "High",
            "Result": "",
            "Defect ID": ""
        },
        {
            "Test ID": "SPI-006",
            "Area": "ADS1293 ECG Acquisition",
            "Objective": "Verify multi-channel ECG data acquisition and streaming",
            "Linked IDs (Hazard/SRS/Task)": "SRS-SPI-006, HAZ-002",
            "Preconditions": "ADS1293 initialized, ECG electrodes connected",
            "Steps": "1. Send JSON start command to port 1293\n2. Configure ECG channels\n3. Start data acquisition\n4. Monitor data quality and continuity",
            "Expected": "Multi-channel ECG data streamed continuously, no data gaps, signal quality acceptable",
            "Dataset": "3-lead ECG",
            "Priority": "Critical",
            "Result": "",
            "Defect ID": ""
        },
        {
            "Test ID": "SPI-007",
            "Area": "ADS1293 Sync Marks",
            "Objective": "Verify ADS1293 sync marks align with MAX30009 sync marks",
            "Linked IDs (Hazard/SRS/Task)": "SRS-SPI-007",
            "Preconditions": "Both MAX30009 and ADS1293 streaming data simultaneously",
            "Steps": "1. Start both sensors simultaneously\n2. Monitor both data streams for 60 seconds\n3. Compare sync mark timestamps\n4. Calculate time drift between sensors",
            "Expected": "Sync marks aligned within ±10ms, sync counter increments match, drift < 100ms over 60 seconds",
            "Dataset": "60 second dual capture",
            "Priority": "Critical",
            "Result": "",
            "Defect ID": ""
        },
        {
            "Test ID": "SPI-008",
            "Area": "WS2812 LED Control",
            "Objective": "Verify WS2812 LED control via TCP port 2812",
            "Linked IDs (Hazard/SRS/Task)": "SRS-SPI-008",
            "Preconditions": "SPI_DEV_servise running, WS2812 LEDs connected",
            "Steps": "1. Connect to port 2812\n2. Send JSON command to set LED color (R=255, G=0, B=0)\n3. Verify LED illuminates red\n4. Test multiple colors and patterns",
            "Expected": "LEDs respond correctly to color commands, no flickering, smooth transitions",
            "Dataset": "RGB test patterns",
            "Priority": "Medium",
            "Result": "",
            "Defect ID": ""
        },
        {
            "Test ID": "SPI-009",
            "Area": "JSON Request/Response",
            "Objective": "Verify JSON request/response mechanism with atomic flags",
            "Linked IDs (Hazard/SRS/Task)": "SRS-SPI-009",
            "Preconditions": "Any TCP server running",
            "Steps": "1. Send malformed JSON to server\n2. Send valid JSON request\n3. Monitor request_ready_flag and response_ready_flag\n4. Verify response JSON format",
            "Expected": "Invalid JSON rejected gracefully, valid JSON processed, atomic flags work correctly, no race conditions",
            "Dataset": "Valid/invalid JSON samples",
            "Priority": "High",
            "Result": "",
            "Defect ID": ""
        },
        {
            "Test ID": "SPI-010",
            "Area": "Main Loop Timing",
            "Objective": "Verify main loop executes with 500μs sleep without timing violations",
            "Linked IDs (Hazard/SRS/Task)": "SRS-SPI-010",
            "Preconditions": "SPI_DEV_servise running with all sensors active",
            "Steps": "1. Start all three sensors simultaneously\n2. Monitor loop execution time with timestamps\n3. Calculate loop period statistics\n4. Check for timing jitter and delays",
            "Expected": "Loop period ~500μs ± 50μs, no missed cycles, consistent timing under load",
            "Dataset": "10,000 loop iterations",
            "Priority": "High",
            "Result": "",
            "Defect ID": ""
        },
        {
            "Test ID": "SPI-011",
            "Area": "Concurrent Sensor Operation",
            "Objective": "Verify all three sensors operate concurrently without interference",
            "Linked IDs (Hazard/SRS/Task)": "SRS-SPI-011, HAZ-003",
            "Preconditions": "All sensors connected and initialized",
            "Steps": "1. Start MAX30009 data acquisition\n2. Start ADS1293 data acquisition\n3. Send WS2812 LED commands\n4. Monitor all data streams simultaneously for 5 minutes\n5. Check for data corruption or dropped samples",
            "Expected": "All sensors operate independently, no data corruption, no resource conflicts, sync marks remain aligned",
            "Dataset": "5 minute concurrent operation",
            "Priority": "Critical",
            "Result": "",
            "Defect ID": ""
        },
        {
            "Test ID": "SPI-012",
            "Area": "SPI Bus Communication",
            "Objective": "Verify SPI communication reliability for MAX30009 and ADS1293",
            "Linked IDs (Hazard/SRS/Task)": "SRS-SPI-012",
            "Preconditions": "SPI hardware drivers initialized",
            "Steps": "1. Read/write test registers on both devices\n2. Verify data integrity with known patterns\n3. Test at maximum SPI clock speed\n4. Monitor for SPI errors in dmesg",
            "Expected": "SPI read/write operations succeed 100%, no bit errors, no CRC failures",
            "Dataset": "1000 read/write cycles",
            "Priority": "High",
            "Result": "",
            "Defect ID": ""
        },
        {
            "Test ID": "SPI-013",
            "Area": "FIFO Buffer Management",
            "Objective": "Verify MAX30009 FIFO buffer handles maximum 30,000 samples without overflow",
            "Linked IDs (Hazard/SRS/Task)": "SRS-SPI-013, HAZ-004",
            "Preconditions": "MAX30009 configured at highest sample rate",
            "Steps": "1. Configure MAX30009 for maximum data rate\n2. Delay reading data to fill FIFO\n3. Monitor FIFO level\n4. Read data before overflow\n5. Verify no samples lost",
            "Expected": "FIFO fills to 30,000 samples maximum, overflow flag triggers before data loss, all samples recovered",
            "Dataset": "Maximum rate test",
            "Priority": "High",
            "Result": "",
            "Defect ID": ""
        },
        {
            "Test ID": "SPI-014",
            "Area": "Service Restart Recovery",
            "Objective": "Verify service gracefully restarts and recovers sensor state",
            "Linked IDs (Hazard/SRS/Task)": "SRS-SPI-014",
            "Preconditions": "Service running with active data acquisition",
            "Steps": "1. Start data acquisition on all sensors\n2. Kill service with SIGTERM\n3. Restart service\n4. Verify sensor re-initialization\n5. Resume data acquisition",
            "Expected": "Service shuts down cleanly, all resources released, restart successful, sensors re-initialize properly",
            "Dataset": "N/A",
            "Priority": "Medium",
            "Result": "",
            "Defect ID": ""
        },
        {
            "Test ID": "SPI-015",
            "Area": "Error Handling",
            "Objective": "Verify error handling for sensor disconnection and SPI failures",
            "Linked IDs (Hazard/SRS/Task)": "SRS-SPI-015, HAZ-005",
            "Preconditions": "Service running normally",
            "Steps": "1. Disconnect MAX30009 sensor during acquisition\n2. Monitor error messages and recovery\n3. Reconnect sensor\n4. Verify automatic recovery or manual restart needed",
            "Expected": "Error detected and logged, service remains stable, clear error message to application, recovery possible",
            "Dataset": "Fault injection test",
            "Priority": "High",
            "Result": "",
            "Defect ID": ""
        }
    ]

    # power_control_service test cases
    pwr_tests = [
        {
            "Test ID": "PWR-001",
            "Area": "Power Control Communication",
            "Objective": "Verify PWRCNTR TCP server accepts connections on port 501",
            "Linked IDs (Hazard/SRS/Task)": "SRS-PWR-001",
            "Preconditions": "power_control_servise running, network available",
            "Steps": "1. Start power_control_servise\n2. Connect to localhost:501 via telnet\n3. Verify connection established",
            "Expected": "TCP connection established successfully, server ready to receive JSON commands",
            "Dataset": "N/A",
            "Priority": "High",
            "Result": "",
            "Defect ID": ""
        },
        {
            "Test ID": "PWR-002",
            "Area": "GPIO Power Control",
            "Objective": "Verify GPIO-based power control switches power domains correctly",
            "Linked IDs (Hazard/SRS/Task)": "SRS-PWR-002, HAZ-006",
            "Preconditions": "libgpiod configured, GPIO pins accessible",
            "Steps": "1. Send JSON command to enable power domain\n2. Verify GPIO pin state using gpio-tools\n3. Measure voltage on power rail\n4. Send disable command\n5. Verify power off",
            "Expected": "GPIO state changes correctly, power rail voltage correct (e.g., 3.3V/5V), power cycles cleanly",
            "Dataset": "Power on/off cycles",
            "Priority": "Critical",
            "Result": "",
            "Defect ID": ""
        },
        {
            "Test ID": "PWR-003",
            "Area": "Battery Monitoring",
            "Objective": "Verify I2C battery status monitoring via SMBus",
            "Linked IDs (Hazard/SRS/Task)": "SRS-PWR-003, HAZ-007",
            "Preconditions": "Battery connected, I2C bus functional",
            "Steps": "1. Send JSON request for battery status\n2. Verify voltage reading accuracy (±50mV)\n3. Verify current reading\n4. Verify state of charge (SOC) percentage\n5. Verify temperature reading",
            "Expected": "Battery voltage accurate, SOC matches expected value, temperature reasonable (15-40°C), data refreshes every 3 seconds",
            "Dataset": "Battery at 50% SOC",
            "Priority": "Critical",
            "Result": "",
            "Defect ID": ""
        },
        {
            "Test ID": "PWR-004",
            "Area": "Battery Reading Throttle",
            "Objective": "Verify battery reading throttled to every 3 seconds",
            "Linked IDs (Hazard/SRS/Task)": "SRS-PWR-004",
            "Preconditions": "Service running, battery monitoring active",
            "Steps": "1. Enable timestamp logging for battery reads\n2. Monitor I2C transactions for 60 seconds\n3. Count number of battery reads\n4. Verify timing between reads",
            "Expected": "Battery read exactly every 3 seconds (±100ms), approximately 20 reads in 60 seconds",
            "Dataset": "60 second monitoring",
            "Priority": "Medium",
            "Result": "",
            "Defect ID": ""
        },
        {
            "Test ID": "PWR-005",
            "Area": "Button Input Handling",
            "Objective": "Verify button input processed correctly with debouncing",
            "Linked IDs (Hazard/SRS/Task)": "SRS-PWR-005",
            "Preconditions": "Button connected to GPIO, service running",
            "Steps": "1. Press button briefly (<100ms)\n2. Press button long (>1s)\n3. Rapidly press button multiple times\n4. Monitor button event reporting via JSON",
            "Expected": "Short press detected, long press detected, no double-triggers from bounce, events reported accurately",
            "Dataset": "Button press patterns",
            "Priority": "High",
            "Result": "",
            "Defect ID": ""
        },
        {
            "Test ID": "PWR-006",
            "Area": "Buzzer Control",
            "Objective": "Verify buzzer control with different patterns and durations",
            "Linked IDs (Hazard/SRS/Task)": "SRS-PWR-006",
            "Preconditions": "Buzzer connected, service running",
            "Steps": "1. Send JSON command for single beep (100ms)\n2. Send command for double beep pattern\n3. Send command for continuous tone (1s)\n4. Test buzzer disable command",
            "Expected": "Buzzer responds accurately to all patterns, timing precise (±10ms), can be interrupted",
            "Dataset": "Buzzer test patterns",
            "Priority": "Medium",
            "Result": "",
            "Defect ID": ""
        },
        {
            "Test ID": "PWR-007",
            "Area": "Main Loop Timing",
            "Objective": "Verify main loop executes with 100ms sleep without timing violations",
            "Linked IDs (Hazard/SRS/Task)": "SRS-PWR-007",
            "Preconditions": "Service running with all features active",
            "Steps": "1. Enable timestamp logging in main loop\n2. Monitor loop execution time for 60 seconds\n3. Calculate loop period statistics\n4. Check for timing jitter",
            "Expected": "Loop period ~100ms ± 10ms, consistent timing, no significant delays",
            "Dataset": "600 loop iterations",
            "Priority": "Medium",
            "Result": "",
            "Defect ID": ""
        },
        {
            "Test ID": "PWR-008",
            "Area": "JSON Request/Response",
            "Objective": "Verify JSON communication with atomic flags and thread safety",
            "Linked IDs (Hazard/SRS/Task)": "SRS-PWR-008",
            "Preconditions": "Service running",
            "Steps": "1. Send concurrent JSON requests from multiple clients\n2. Send malformed JSON\n3. Send valid requests rapidly\n4. Monitor request/response flags",
            "Expected": "Thread-safe operation, no race conditions, invalid JSON rejected, all valid requests processed in order",
            "Dataset": "Concurrent requests",
            "Priority": "High",
            "Result": "",
            "Defect ID": ""
        },
        {
            "Test ID": "PWR-009",
            "Area": "I2C Communication Reliability",
            "Objective": "Verify SMBus/I2C communication handles errors gracefully",
            "Linked IDs (Hazard/SRS/Task)": "SRS-PWR-009, HAZ-008",
            "Preconditions": "I2C bus functional",
            "Steps": "1. Disconnect battery during active monitoring\n2. Monitor I2C error handling\n3. Reconnect battery\n4. Verify recovery\n5. Check error logging",
            "Expected": "I2C timeout detected, error logged, service remains stable, automatic recovery when battery reconnected",
            "Dataset": "Fault injection",
            "Priority": "High",
            "Result": "",
            "Defect ID": ""
        },
        {
            "Test ID": "PWR-010",
            "Area": "Low Battery Handling",
            "Objective": "Verify low battery detection and notification",
            "Linked IDs (Hazard/SRS/Task)": "SRS-PWR-010, HAZ-009",
            "Preconditions": "Battery at <10% SOC or use battery simulator",
            "Steps": "1. Monitor battery SOC\n2. Drain battery to <10%\n3. Verify low battery notification sent\n4. Check critical battery threshold (<5%)\n5. Verify shutdown sequence initiated",
            "Expected": "Low battery warning at 10%, critical warning at 5%, notification sent via JSON, graceful shutdown initiated",
            "Dataset": "Low battery scenario",
            "Priority": "Critical",
            "Result": "",
            "Defect ID": ""
        },
        {
            "Test ID": "PWR-011",
            "Area": "Power Sequencing",
            "Objective": "Verify correct power-up and power-down sequencing",
            "Linked IDs (Hazard/SRS/Task)": "SRS-PWR-011, HAZ-010",
            "Preconditions": "All power domains available",
            "Steps": "1. Send power-up command\n2. Monitor GPIO timing with oscilloscope\n3. Verify power rail sequencing\n4. Send power-down command\n5. Verify reverse sequencing",
            "Expected": "Power rails enable in correct order with proper delays, power-down sequence is reverse of power-up, no voltage spikes",
            "Dataset": "Oscilloscope capture",
            "Priority": "Critical",
            "Result": "",
            "Defect ID": ""
        },
        {
            "Test ID": "PWR-012",
            "Area": "Concurrent Operations",
            "Objective": "Verify button, buzzer, and battery monitoring work concurrently",
            "Linked IDs (Hazard/SRS/Task)": "SRS-PWR-012",
            "Preconditions": "All peripherals connected",
            "Steps": "1. Start battery monitoring\n2. Press button while buzzer active\n3. Monitor all features simultaneously for 5 minutes\n4. Verify no interference",
            "Expected": "All features work independently, no missed button presses, battery reads continue, buzzer unaffected",
            "Dataset": "5 minute test",
            "Priority": "High",
            "Result": "",
            "Defect ID": ""
        },
        {
            "Test ID": "PWR-013",
            "Area": "Service Restart Recovery",
            "Objective": "Verify service restarts cleanly and restores power state",
            "Linked IDs (Hazard/SRS/Task)": "SRS-PWR-013",
            "Preconditions": "Service running with power enabled",
            "Steps": "1. Enable power domains\n2. Kill service with SIGTERM\n3. Verify power state maintained by hardware\n4. Restart service\n5. Verify state synchronization",
            "Expected": "Service shuts down cleanly, power state preserved or safely restored, no GPIO conflicts on restart",
            "Dataset": "N/A",
            "Priority": "High",
            "Result": "",
            "Defect ID": ""
        },
        {
            "Test ID": "PWR-014",
            "Area": "Temperature Monitoring",
            "Objective": "Verify battery temperature monitoring and over-temperature protection",
            "Linked IDs (Hazard/SRS/Task)": "SRS-PWR-014, HAZ-011",
            "Preconditions": "Battery with temperature sensor",
            "Steps": "1. Monitor normal temperature (20-30°C)\n2. Heat battery to 45°C (use controlled environment)\n3. Verify over-temperature warning\n4. Heat to 50°C\n5. Verify shutdown initiated",
            "Expected": "Temperature read accurately (±2°C), warning at 45°C, shutdown at 50°C, notification sent to application",
            "Dataset": "Temperature sweep",
            "Priority": "Critical",
            "Result": "",
            "Defect ID": ""
        },
        {
            "Test ID": "PWR-015",
            "Area": "Inter-Service Communication",
            "Objective": "Verify power_control_servise can coordinate with SPI_DEV_servise",
            "Linked IDs (Hazard/SRS/Task)": "SRS-PWR-015, SRS-SPI-016",
            "Preconditions": "Both services running",
            "Steps": "1. Send power enable to power_control_servise\n2. Verify SPI_DEV_servise detects power and initializes sensors\n3. Send low battery command\n4. Verify SPI_DEV_servise receives notification and saves data\n5. Coordinate shutdown",
            "Expected": "Services communicate through application layer, coordinated startup/shutdown, data preserved on low battery",
            "Dataset": "Integration test",
            "Priority": "High",
            "Result": "",
            "Defect ID": ""
        }
    ]

    # Write headers and test cases to both sheets
    for ws, tests in [(ws_spi, spi_tests), (ws_pwr, pwr_tests)]:
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Write test cases
        for row_num, test in enumerate(tests, 2):
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num, value=test[header])
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border

        # Adjust column widths
        column_widths = {
            'A': 12,  # Test ID
            'B': 25,  # Area
            'C': 35,  # Objective
            'D': 20,  # Linked IDs
            'E': 30,  # Preconditions
            'F': 50,  # Steps
            'G': 45,  # Expected
            'H': 20,  # Dataset
            'I': 10,  # Priority
            'J': 10,  # Result
            'K': 12   # Defect ID
        }

        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Set row height for header
        ws.row_dimensions[1].height = 30

        # Auto-fit row heights for content
        for row in range(2, len(tests) + 2):
            ws.row_dimensions[row].height = None  # Auto

    # Save workbook
    filename = f"Firmware_Test_Cases_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    print(f"Test cases generated successfully: {filename}")
    return filename

if __name__ == "__main__":
    create_test_cases_excel()
